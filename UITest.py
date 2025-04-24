import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from functools import partial
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border
import math
import os
import sys
from openpyxl import load_workbook
import threading

def select_import_file():
    file_path = filedialog.askopenfilename(
    title="Select a file",
    initialdir="C:\\",
    filetypes=(("Excel Files", "*.xlsx"), ("All files", "*.*"))
    )
    file_to_import.set(file_path)
    if(folder_to_export.get() == ""):
        print(f"Output dir: {"/".join(file_path.split("/")[:-1])}")
        folder_to_export.set("/".join(file_path.split("/")[:-1]))

def select_output_dir():
    file_path = filedialog.askdirectory(
    title="Select a directory",
    initialdir="C:\\",
    )
    folder_to_export.set(file_path)

def progress_popup():
    pass

def print_help():
    print("This program takes a file as input for the first argument.")
    print("Optionally, include a value as the second argument for how many entries needed per file.")
    print("use the argument -help to see this dialogue.")
    sys.exit()

def split():
    import_file = file_to_import.get()
    output_dir = folder_to_export.get()

    #define max rows
    max_rows_per_file = split_length.get()
    
    #zero out progress bar
    progress.set(1)

        
    #set the export file name
    print(f"Import File: {import_file}")
    export_file = import_file.split("/")
    for s in export_file:
        if ".xlsx" in s:
            export_file = s.split(".")[0]
    print(f"*******Export File: {export_file}")
    #make output directory
    os.makedirs(f"{output_dir}/output", exist_ok=True)
    output_dir = f"{output_dir}/output/"

    #load the workbook and set the active worksheet
    try:
        wb = load_workbook(filename = import_file, rich_text = True)
        ws = wb.active
    except Exception as e:
        print(f"An error occured: {e}")
        sys.exit()
    
    



  

    #calculate number of files needed
    files_needed = math.ceil((ws.max_row) / max_rows_per_file)
    print(f"splitting {import_file} into {files_needed} files")
    for i in range(files_needed):
        #update progress bar
        if((i/files_needed) *100 != 0):
            progress.set((i/files_needed) *100)
        print(f"progress: {progress}, {i/files_needed *100}")
        #create new workbook to store new data
        wb_out = Workbook()
        ws_out = wb_out.active

        #copy rows to newely created workbook
        for row in ws.iter_rows(min_row=((i * max_rows_per_file)+1), max_row = ((i+1) * max_rows_per_file) if (((i+1) * max_rows_per_file) < ws.max_row) else ws.max_row):
            new_row = []
            #add leading zeroes for first column
            for cell in row:
                if(cell.column == 1 and isinstance(cell.value, int)):
                    new_row.append(f'{cell.value:010d}')
                #if not first column, add manually    
                else:
                    new_row.append(cell.value)
            if(cell.row != 1):
                ws_out.append(new_row)
        print(f"Saving {export_file} {i+1}.xlsx to {output_dir}{export_file} {i+1}.xlsx")
        wb_out.save(f"{output_dir}{export_file} {i+1}.xlsx")
        i+=1
    #set progress to 100
    progress.set(100)
    #display completion popup
    messagebox.showinfo("Status", "Excel file split successfully.")

def startSplit():
    threading.Thread(target=split).start()

##############Define UI##############
window = tk.Tk()
window.title("Excel Splitter")

mainframe = ttk.Frame(window, padding="3 3 12 12")
mainframe.grid(column=2, row=7)
window.columnconfigure(0, weight=1)
window.rowconfigure(0, weight=1)
file_to_import = tk.StringVar()
folder_to_export = tk.StringVar()
progress = tk.IntVar()
split_length = tk.IntVar()
split_length.set(60000)
trim_headers = tk.BooleanVar()

#import file UI
ttk.Label(mainframe, text="File to split:").grid(row=1, column=1)
import_entry = ttk.Entry(mainframe, width=100, textvariable=file_to_import)
import_entry.grid(column=1, row=2, sticky=tk.W)
ttk.Button(mainframe, text="open", command=select_import_file).grid(row=2, column=2)

#export file UI
ttk.Label(mainframe, text="Output Directory:").grid(row=3, column=1)
export_entry = ttk.Entry(mainframe, width=100, textvariable=folder_to_export)
export_entry.grid(column=1, row=4, sticky=tk.W)
ttk.Button(mainframe, text="open", command=select_output_dir).grid(row=4, column=2)

#Split length UI
length_containter = ttk.Frame(mainframe)
length_containter.grid(row=5, column=1)
ttk.Label(length_containter, text="Lines per file:").grid(row=0, column=0, sticky=tk.W)
length_entry = ttk.Entry(length_containter, width=25, textvariable=split_length).grid(row=0, column=1, sticky=tk.W)


#Trim Header UI
ttk.Checkbutton(mainframe, text='Trim Header', variable=trim_headers).grid(row=5, column=2)

#Start button UI
ttk.Button(mainframe, text="Split", command=startSplit).grid(row=6, column=1)

#progress bar UI
progress_bar = ttk.Progressbar(mainframe, orient="horizontal", length=500, mode="determinate", maximum=100, variable=progress).grid(row=7, column=1)

window.mainloop()
